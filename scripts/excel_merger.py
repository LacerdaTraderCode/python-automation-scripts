"""
Excel Merger - combina múltiplas planilhas Excel em um único arquivo.

Uso: python scripts/excel_merger.py <pasta_com_xlsx> <arquivo_saida.xlsx>
"""
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook


def merge_excels(input_dir: str, output_file: str):
    """Combina todas as planilhas .xlsx de uma pasta."""
    input_path = Path(input_dir).expanduser().resolve()
    output_path = Path(output_file).expanduser().resolve()

    if not input_path.is_dir():
        print(f"❌ Pasta inválida: {input_path}")
        return

    excel_files = list(input_path.glob("*.xlsx"))
    if not excel_files:
        print(f"❌ Nenhum arquivo .xlsx encontrado em {input_path}")
        return

    print(f"📊 Combinando {len(excel_files)} arquivo(s)...\n")

    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # Remove sheet padrão

    total_rows = 0
    for excel_file in sorted(excel_files):
        print(f"  📄 {excel_file.name}")
        wb = load_workbook(excel_file, data_only=True)

        for sheet_name in wb.sheetnames:
            source_sheet = wb[sheet_name]
            # Nome único: arquivo + aba
            new_name = f"{excel_file.stem}_{sheet_name}"[:31]  # Excel limita 31 chars
            new_sheet = merged_wb.create_sheet(title=new_name)

            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)
                total_rows += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged_wb.save(output_path)

    print(f"\n✅ Combinado!")
    print(f"   Arquivos processados: {len(excel_files)}")
    print(f"   Abas criadas: {len(merged_wb.sheetnames)}")
    print(f"   Linhas totais: {total_rows:,}")
    print(f"   Saída: {output_path}")


def main():
    if len(sys.argv) < 3:
        print("Uso: python scripts/excel_merger.py <pasta_com_xlsx> <arquivo_saida.xlsx>")
        sys.exit(1)

    merge_excels(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
